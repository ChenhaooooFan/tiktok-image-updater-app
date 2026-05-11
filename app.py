import csv
import io
import re
import unicodedata
from collections import defaultdict
from pathlib import PurePosixPath
from urllib.parse import quote
from zipfile import ZipFile, ZIP_DEFLATED

import streamlit as st
from openpyxl import load_workbook


st.set_page_config(
    page_title="NailVesta TikTok Image Batch Updater",
    page_icon="💅",
    layout="wide",
)


# =========================
# Basic helpers
# =========================

VALID_EXTS = {".jpg", ".jpeg", ".png"}


def remove_accent(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c))


def parse_aliases(alias_text: str) -> dict:
    aliases = {
        "rosey tigress": "rosy tigress",
        "rosy tigress": "rosy tigress",
        "rose angel": "rose angel",
        "rose petal": "rose petal",
    }

    for line in alias_text.splitlines():
        line = line.strip()
        if not line or "=" not in line:
            continue

        left, right = line.split("=", 1)
        left = left.strip().lower()
        right = right.strip().lower()

        if left and right:
            aliases[left] = right

    return aliases


def normalize_style(name: str, aliases: dict) -> str:
    if not name:
        return ""

    name = str(name).strip()
    name = remove_accent(name)

    name = name.replace("’", "'")
    name = name.replace("‘", "'")
    name = name.replace("`", "'")
    name = name.replace("–", "-")
    name = name.replace("—", "-")
    name = name.replace("-", " ")

    name = re.sub(r"\s+", " ", name)
    name = name.lower().strip()

    return aliases.get(name, name)


def extract_style_from_variation(variation_value: str) -> str:
    if not variation_value:
        return ""

    return str(variation_value).split(",")[0].strip()


def extract_style_from_filename(filename_stem: str):
    match = re.match(r"^(.*)_(\d+)$", filename_stem)
    if not match:
        return None, None

    style_name = match.group(1).strip()
    image_number = int(match.group(2))
    return style_name, image_number


def make_public_url(public_base: str, folder_name: str, filename: str) -> str:
    public_base = public_base.rstrip("/")
    folder_name = folder_name.strip("/")

    return f"{public_base}/{folder_name}/{quote(filename)}"


def get_col_by_header(ws, header_name: str):
    """
    Search first 5 rows, because TikTok template sometimes has visible guide rows.
    Prefer exact system header like main_image / image_2 / variation_value.
    """
    for row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if value and str(value).strip() == header_name:
                return row, col

    raise ValueError(f"找不到列: {header_name}")


# =========================
# XLSX repair
# =========================

def repair_xlsx_bytes(raw_bytes: bytes) -> bytes:
    """
    Removes broken styles / conditional formatting from TikTok workbook.
    This preserves worksheet data and URLs.
    """
    minimal_styles = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1">
    <font>
      <sz val="11"/>
      <color theme="1"/>
      <name val="Calibri"/>
      <family val="2"/>
      <scheme val="minor"/>
    </font>
  </fonts>
  <fills count="2">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
  </fills>
  <borders count="1">
    <border>
      <left/><right/><top/><bottom/><diagonal/>
    </border>
  </borders>
  <cellStyleXfs count="1">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>
  </cellStyleXfs>
  <cellXfs count="1">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
  </cellXfs>
  <cellStyles count="1">
    <cellStyle name="Normal" xfId="0" builtinId="0"/>
  </cellStyles>
  <dxfs count="0"/>
  <tableStyles count="0" defaultTableStyle="TableStyleMedium9" defaultPivotStyle="PivotStyleLight16"/>
</styleSheet>
"""

    src = io.BytesIO(raw_bytes)
    dst = io.BytesIO()

    with ZipFile(src, "r") as zin, ZipFile(dst, "w", ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)

            if item.filename == "xl/styles.xml":
                zout.writestr(item, minimal_styles)
                continue

            if item.filename == "xl/calcChain.xml":
                continue

            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                text = data.decode("utf-8")

                text = re.sub(r'\s+s="\d+"', "", text)

                text = re.sub(
                    r"<conditionalFormatting\b[^>]*>.*?</conditionalFormatting>",
                    "",
                    text,
                    flags=re.DOTALL,
                )

                text = re.sub(
                    r"<conditionalFormatting\b[^>]*/>",
                    "",
                    text,
                    flags=re.DOTALL,
                )

                text = re.sub(
                    r"<extLst\b[^>]*>.*?</extLst>",
                    "",
                    text,
                    flags=re.DOTALL,
                )

                data = text.encode("utf-8")

            if item.filename.startswith("xl/_rels/workbook.xml.rels"):
                text = data.decode("utf-8")
                text = re.sub(
                    r'<Relationship[^>]+calcChain[^>]*/>',
                    "",
                    text,
                )
                data = text.encode("utf-8")

            zout.writestr(item, data)

    return dst.getvalue()


def load_workbook_safely(raw_bytes: bytes):
    try:
        wb = load_workbook(io.BytesIO(raw_bytes))
        return wb, raw_bytes, False
    except Exception:
        repaired = repair_xlsx_bytes(raw_bytes)
        wb = load_workbook(io.BytesIO(repaired))
        return wb, repaired, True


# =========================
# ZIP image readers
# =========================

def list_images_from_zip(uploaded_zip, expected_type: str):
    """
    Returns image base filenames from uploaded ZIP.
    Ignores folders, __MACOSX, hidden files.
    """
    image_names = []

    with ZipFile(uploaded_zip, "r") as z:
        for name in z.namelist():
            if name.endswith("/"):
                continue

            path = PurePosixPath(name)
            filename = path.name

            if not filename:
                continue

            if filename.startswith(".") or "__MACOSX" in name:
                continue

            suffix = PurePosixPath(filename).suffix.lower()
            if suffix not in VALID_EXTS:
                continue

            image_names.append(filename)

    image_names = sorted(set(image_names), key=lambda x: x.lower())

    if expected_type == "common" and len(image_names) != 6:
        raise ValueError(f"Common Images 必须正好 6 张，目前找到 {len(image_names)} 张。")

    return image_names


def build_style_map(style_filenames, public_base, style_folder_name, aliases):
    style_map = defaultdict(list)
    bad_filename_images = []

    for filename in style_filenames:
        stem = PurePosixPath(filename).stem
        style_name, image_number = extract_style_from_filename(stem)

        if not style_name:
            bad_filename_images.append(filename)
            continue

        key = normalize_style(style_name, aliases)

        style_map[key].append({
            "style_name_original": style_name,
            "normalized_style": key,
            "number": image_number,
            "filename": filename,
            "url": make_public_url(public_base, style_folder_name, filename),
        })

    for key in style_map:
        style_map[key] = sorted(style_map[key], key=lambda x: x["number"])

    return dict(style_map), bad_filename_images



def list_images_from_txt(uploaded_txt, expected_type: str):
    content = uploaded_txt.getvalue().decode("utf-8-sig")
    image_names = []

    for line in content.splitlines():
        filename = line.strip()
        if not filename:
            continue

        suffix = PurePosixPath(filename).suffix.lower()
        if suffix not in VALID_EXTS:
            continue

        image_names.append(filename)

    image_names = sorted(set(image_names), key=lambda x: x.lower())

    if expected_type == "common" and len(image_names) != 6:
        raise ValueError(f"Common Images 必须正好 6 张，目前找到 {len(image_names)} 张。")

    return image_names



def read_catalog_styles_from_csv(uploaded_csv, aliases: dict):
    raw = uploaded_csv.getvalue()

    text = None
    for enc in ["utf-8-sig", "utf-8", "gb18030"]:
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue

    if text is None:
        raise ValueError("产品图册 CSV 编码无法读取，请另存为 UTF-8 CSV 后再上传。")

    reader = csv.DictReader(io.StringIO(text))

    if not reader.fieldnames:
        raise ValueError("产品图册 CSV 没有表头。")

    clean_fieldnames = [str(x).strip() for x in reader.fieldnames]

    preferred_columns = [
        "款式英文名称",
        "款式英文名",
        "Style English Name",
        "English Style Name",
        "Style Name",
        "style_name",
        "选项引用列",
    ]

    style_col = None

    for col in preferred_columns:
        if col in clean_fieldnames:
            style_col = col
            break

    if style_col is None:
        for col in clean_fieldnames:
            lower_col = col.lower()
            if ("款式" in col and "英文" in col) or ("style" in lower_col and "name" in lower_col):
                style_col = col
                break

    if style_col is None:
        raise ValueError(
            "找不到产品图册里的款式英文名列。请确认列名是否为：款式英文名称"
        )

    catalog_styles = {}

    for row in reader:
        raw_style = row.get(style_col)
        if not raw_style:
            continue

        raw_style = str(raw_style).strip()
        key = normalize_style(raw_style, aliases)

        if key:
            catalog_styles.setdefault(key, raw_style)

    return catalog_styles, style_col


def split_missing_by_catalog(missing_unique: dict, catalog_styles: dict):
    missing_in_catalog = []
    missing_not_in_catalog = []

    for key, tiktok_style in sorted(missing_unique.items()):
        row = {
            "Normalized style": key,
            "Style name from TikTok": tiktok_style,
        }

        if key in catalog_styles:
            row["Style name from catalog"] = catalog_styles[key]
            missing_in_catalog.append(row)
        else:
            missing_not_in_catalog.append(row)

    return missing_in_catalog, missing_not_in_catalog


def make_catalog_compare_csv(missing_in_catalog, missing_not_in_catalog, catalog_column):
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["PRODUCT CATALOG COMPARE"])
    writer.writerow(["Catalog style column", catalog_column])
    writer.writerow([])

    writer.writerow(["NOT IN PRODUCT CATALOG AND ALSO UNMATCHED"])
    writer.writerow(["Normalized style", "Style name from TikTok"])
    for row in missing_not_in_catalog:
        writer.writerow([
            row.get("Normalized style"),
            row.get("Style name from TikTok"),
        ])

    writer.writerow([])

    writer.writerow(["IN PRODUCT CATALOG BUT UNMATCHED"])
    writer.writerow(["Normalized style", "Style name from TikTok", "Style name from catalog"])
    for row in missing_in_catalog:
        writer.writerow([
            row.get("Normalized style"),
            row.get("Style name from TikTok"),
            row.get("Style name from catalog"),
        ])

    return output.getvalue().encode("utf-8-sig")


# =========================
# Main Excel processing
# =========================

def update_template(
    wb,
    style_map,
    common_urls,
    data_start_row: int,
):
    ws = wb.active

    header_row, col_variation = get_col_by_header(ws, "variation_value")

    _, col_main_image = get_col_by_header(ws, "main_image")
    _, col_image_2 = get_col_by_header(ws, "image_2")
    _, col_image_3 = get_col_by_header(ws, "image_3")
    _, col_image_4 = get_col_by_header(ws, "image_4")
    _, col_image_5 = get_col_by_header(ws, "image_5")
    _, col_image_6 = get_col_by_header(ws, "image_6")
    _, col_image_7 = get_col_by_header(ws, "image_7")
    _, col_image_8 = get_col_by_header(ws, "image_8")
    _, col_image_9 = get_col_by_header(ws, "image_9")

    image_cols_common = [
        col_image_4,
        col_image_5,
        col_image_6,
        col_image_7,
        col_image_8,
        col_image_9,
    ]

    updated_rows = 0
    skipped_rows = 0

    missing_rows = []
    missing_unique = {}

    less_than_two_rows = []
    less_than_two_unique = {}

    used_style_keys = set()

    for row in range(data_start_row, ws.max_row + 1):
        variation_value = ws.cell(row=row, column=col_variation).value

        if not variation_value:
            skipped_rows += 1
            continue

        style_name = extract_style_from_variation(variation_value)
        key = normalize_style(style_name, st.session_state.aliases)

        if not key:
            skipped_rows += 1
            continue

        if key not in style_map:
            # Keep old images unchanged.
            missing_rows.append([row, style_name, variation_value])
            missing_unique[key] = style_name
            continue

        images = style_map[key]

        if len(images) < 2:
            # Keep old images unchanged.
            less_than_two_rows.append([row, style_name, len(images), variation_value])
            less_than_two_unique[key] = [style_name, len(images)]
            continue

        used_style_keys.add(key)

        # Main image is never touched.
        ws.cell(row=row, column=col_image_2).value = images[0]["url"]
        ws.cell(row=row, column=col_image_3).value = images[1]["url"]

        for col, url in zip(image_cols_common, common_urls):
            ws.cell(row=row, column=col).value = url

        updated_rows += 1

    unused_style_groups = []
    for key, images in style_map.items():
        if key not in used_style_keys:
            unused_style_groups.append([
                images[0]["style_name_original"],
                key,
                ", ".join(img["filename"] for img in images),
            ])

    result = {
        "header_row": header_row,
        "data_start_row": data_start_row,
        "updated_rows": updated_rows,
        "skipped_rows": skipped_rows,
        "missing_rows": missing_rows,
        "missing_unique": missing_unique,
        "less_than_two_rows": less_than_two_rows,
        "less_than_two_unique": less_than_two_unique,
        "unused_style_groups": unused_style_groups,
        "columns": {
            "main_image": col_main_image,
            "image_2": col_image_2,
            "image_3": col_image_3,
            "image_4": col_image_4,
            "image_5": col_image_5,
            "image_6": col_image_6,
            "image_7": col_image_7,
            "image_8": col_image_8,
            "image_9": col_image_9,
        }
    }

    return wb, result


def workbook_to_bytes(wb):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def make_report_csv(result, style_map, bad_filename_images, common_filenames):
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["SUMMARY"])
    writer.writerow(["Updated rows", result["updated_rows"]])
    writer.writerow(["Skipped blank rows", result["skipped_rows"]])
    writer.writerow(["Missing rows count", len(result["missing_rows"])])
    writer.writerow(["Missing unique styles count", len(result["missing_unique"])])
    writer.writerow(["Rows with less than 2 images", len(result["less_than_two_rows"])])
    writer.writerow(["Unique styles with less than 2 images", len(result["less_than_two_unique"])])
    writer.writerow(["Unused style image groups", len(result["unused_style_groups"])])
    writer.writerow(["Bad filename images", len(bad_filename_images)])
    writer.writerow(["Style image groups loaded", len(style_map)])
    writer.writerow(["Common images loaded", len(common_filenames)])
    writer.writerow(["Header row detected", result["header_row"]])
    writer.writerow(["Data start row used", result["data_start_row"]])

    for name, col in result["columns"].items():
        writer.writerow([f"{name} column", col])

    writer.writerow([])

    writer.writerow(["COMMON IMAGES USED"])
    writer.writerow(["Image slot", "Filename"])
    for i, filename in enumerate(common_filenames, start=4):
        writer.writerow([f"image_{i}", filename])

    writer.writerow([])

    writer.writerow(["MISSING UNIQUE STYLES"])
    writer.writerow(["Normalized style", "Style name from TikTok"])
    for key, style_name in sorted(result["missing_unique"].items()):
        writer.writerow([key, style_name])

    writer.writerow([])

    writer.writerow(["MISSING ROW DETAILS"])
    writer.writerow(["Excel row", "Style name from TikTok", "Original variation value"])
    writer.writerows(result["missing_rows"])

    writer.writerow([])

    writer.writerow(["STYLES WITH LESS THAN 2 IMAGES"])
    writer.writerow(["Excel row", "Style name", "Image count", "Original variation value"])
    writer.writerows(result["less_than_two_rows"])

    writer.writerow([])

    writer.writerow(["UNUSED STYLE IMAGE GROUPS"])
    writer.writerow(["Style name from image file", "Normalized style", "Image files"])
    writer.writerows(result["unused_style_groups"])

    writer.writerow([])

    writer.writerow(["BAD FILENAME IMAGES"])
    writer.writerow(["Filename"])
    for name in bad_filename_images:
        writer.writerow([name])

    return output.getvalue().encode("utf-8-sig")


def make_link_check_csv(wb):
    ws = wb.active

    _, col_product_name = get_col_by_header(ws, "product_name")
    _, col_variation = get_col_by_header(ws, "variation_value")
    _, col_main_image = get_col_by_header(ws, "main_image")
    _, col_image_2 = get_col_by_header(ws, "image_2")
    _, col_image_3 = get_col_by_header(ws, "image_3")
    _, col_image_4 = get_col_by_header(ws, "image_4")
    _, col_image_5 = get_col_by_header(ws, "image_5")
    _, col_image_6 = get_col_by_header(ws, "image_6")
    _, col_image_7 = get_col_by_header(ws, "image_7")
    _, col_image_8 = get_col_by_header(ws, "image_8")
    _, col_image_9 = get_col_by_header(ws, "image_9")

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Excel Row",
        "Product Name",
        "Variation",
        "Main Image",
        "Image 2",
        "Image 3",
        "Image 4",
        "Image 5",
        "Image 6",
        "Image 7",
        "Image 8",
        "Image 9",
    ])

    for row in range(6, ws.max_row + 1):
        product_name = ws.cell(row=row, column=col_product_name).value
        variation = ws.cell(row=row, column=col_variation).value

        if not product_name and not variation:
            continue

        writer.writerow([
            row,
            product_name,
            variation,
            ws.cell(row=row, column=col_main_image).value,
            ws.cell(row=row, column=col_image_2).value,
            ws.cell(row=row, column=col_image_3).value,
            ws.cell(row=row, column=col_image_4).value,
            ws.cell(row=row, column=col_image_5).value,
            ws.cell(row=row, column=col_image_6).value,
            ws.cell(row=row, column=col_image_7).value,
            ws.cell(row=row, column=col_image_8).value,
            ws.cell(row=row, column=col_image_9).value,
        ])

    return output.getvalue().encode("utf-8-sig")


# =========================
# Streamlit UI
# =========================

st.title("💅 NailVesta TikTok 图片批量更新工具")

st.caption(
    "主图 main_image 不会被修改；只更新 Image 2 / Image 3 和 Image 4-9。"
)

with st.sidebar:
    st.header("Supabase 设置")

    public_base = st.text_input(
        "Public base URL",
        value="https://uuhbjnqrzasegwawhjht.supabase.co/storage/v1/object/public/tiktok-product-images",
    )

    style_folder_name = st.text_input("Style folder", value="style")
    common_folder_name = st.text_input("Common folder", value="common")

    data_start_row = st.number_input(
        "TikTok 产品数据开始行",
        min_value=2,
        max_value=20,
        value=6,
        step=1,
    )

    st.header("款式别名")
    alias_text = st.text_area(
        "一行一个：TikTok/图片里的写法 = 统一写法",
        value="rosey tigress=rosy tigress\n",
        height=120,
    )

st.session_state.aliases = parse_aliases(alias_text)

template_file = st.file_uploader(
    "1. 上传 TikTok 后台下载的模板 .xlsx",
    type=["xlsx"],
)

style_txt = st.file_uploader(
    "2. 上传 style_filenames.txt",
    type=["txt"],
)

common_txt = st.file_uploader(
    "3. 上传 common_filenames.txt",
    type=["txt"],
)

product_catalog_file = st.file_uploader(
    "4. 上传产品图册 CSV（用于判断未匹配款式是否在图册里）",
    type=["csv"],
)

run = st.button("生成 TikTok 上传文件", type="primary")

if run:
    if not template_file or not style_txt or not common_txt or not product_catalog_file:
        st.error("请先上传 TikTok 模板、style_filenames.txt、common_filenames.txt、产品图册 CSV。")
        st.stop()

    try:
        with st.spinner("读取并修复 TikTok 模板..."):
            raw_template = template_file.getvalue()
            wb, repaired_bytes, was_repaired = load_workbook_safely(raw_template)

        if was_repaired:
            st.warning("TikTok 模板有样式问题，已自动修复后继续处理。")
        else:
            st.success("TikTok 模板读取正常。")

        with st.spinner("读取图片文件名列表..."):
            style_filenames = list_images_from_txt(style_txt, expected_type="style")
            common_filenames = list_images_from_txt(common_txt, expected_type="common")

        aliases = st.session_state.aliases
        catalog_styles, catalog_column = read_catalog_styles_from_csv(product_catalog_file, aliases)

        style_map, bad_filename_images = build_style_map(
            style_filenames,
            public_base,
            style_folder_name,
            aliases,
        )

        common_urls = [
            make_public_url(public_base, common_folder_name, filename)
            for filename in common_filenames
        ]

        with st.spinner("生成 TikTok 上传 Excel..."):
            updated_wb, result = update_template(
                wb,
                style_map,
                common_urls,
                data_start_row=int(data_start_row),
            )

            final_xlsx = workbook_to_bytes(updated_wb)
            report_csv = make_report_csv(
                result,
                style_map,
                bad_filename_images,
                common_filenames,
            )
            link_check_csv = make_link_check_csv(updated_wb)

            missing_in_catalog, missing_not_in_catalog = split_missing_by_catalog(
                result["missing_unique"],
                catalog_styles,
            )

            catalog_compare_csv = make_catalog_compare_csv(
                missing_in_catalog,
                missing_not_in_catalog,
                catalog_column,
            )

        st.success("处理完成。")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("成功更新行数", result["updated_rows"])
        c2.metric("未匹配行数", len(result["missing_rows"]))
        c3.metric("未匹配唯一款式", len(result["missing_unique"]))
        c4.metric("少于2张图", len(result["less_than_two_rows"]))

        st.subheader("识别到的图片列")
        st.write(result["columns"])

        if result["columns"].get("main_image") != 20:
            st.warning(
                "提示：这次 main_image 不在第 20 列。"
                "不一定是错误，因为 app 会按字段名写入；但建议你抽查上传文件。"
            )

        st.subheader("下载文件")

        st.download_button(
            label="下载 TikTok 上传 Excel",
            data=final_xlsx,
            file_name="TikTok_Bulk_Image_Update_Final_FIXED.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.download_button(
            label="下载检查报告 CSV",
            data=report_csv,
            file_name="TikTok_Image_Update_Report_FIXED.csv",
            mime="text/csv",
        )

        st.download_button(
            label="下载可复制链接检查表 CSV",
            data=link_check_csv,
            file_name="Image_Link_Check_Report.csv",
            mime="text/csv",
        )

        st.download_button(
            label="下载未匹配款式 × 产品图册对照 CSV",
            data=catalog_compare_csv,
            file_name="Missing_Styles_Catalog_Compare.csv",
            mime="text/csv",
        )

        st.subheader("未匹配款式 × 产品图册对照")

        st.caption(f"产品图册英文款式列：{catalog_column}")

        c5, c6 = st.columns(2)
        c5.metric("不在产品图册也未匹配", len(missing_not_in_catalog))
        c6.metric("在产品图册但未匹配", len(missing_in_catalog))

        st.write("### 不在产品图册，也未匹配")
        if missing_not_in_catalog:
            st.dataframe(missing_not_in_catalog, use_container_width=True)
        else:
            st.success("没有这类款式。")

        st.write("### 在产品图册，但未匹配")
        if missing_in_catalog:
            st.dataframe(missing_in_catalog, use_container_width=True)
        else:
            st.success("没有这类款式。")

        st.info(
            "上传 TikTok 的文件是 TikTok_Bulk_Image_Update_Final_FIXED.xlsx。"
            "CSV 只是检查用，不要上传 TikTok。"
        )

    except Exception as e:
        st.error("处理失败。")
        st.exception(e)
